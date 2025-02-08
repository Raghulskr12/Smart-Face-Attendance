import cv2  # OpenCV library for computer vision tasks
import numpy as np  # NumPy library for numerical operations
import face_recognition  # face_recognition library for face detection and recognition
import os  # OS library for interacting with the operating system
from datetime import datetime  # datetime library for handling dates and times
import openpyxl  # openpyxl library for working with Excel files

# Constants
TRAINING_IMAGES_PATH = 'Training_images'  # Path to the directory containing training images
ATTENDANCE_EXCEL_FILE = 'Attendance.xlsx'  # Name of the Excel file to store attendance
FACE_RECOGNITION_THRESHOLD = 0.5  # Adjust this threshold as needed. Lower values make recognition stricter

def load_training_images(path):
    """Load training images and corresponding class names."""
    images = []  # Initialize an empty list to store the images
    class_names = []  # Initialize an empty list to store the corresponding class names
    for image_file in os.listdir(path):  # Iterate over each file in the specified directory
        image_path = os.path.join(path, image_file)  # Create the full path to the image file
        image = cv2.imread(image_path)  # Read the image using OpenCV
        if image is not None:  # Check if the image was successfully loaded
            images.append(image)  # Add the image to the list of images
            class_names.append(os.path.splitext(image_file)[0])  # Extract the class name (filename without extension) and add it to the list
    return images, class_names  # Return the list of images and the corresponding class names

def preprocess_image(img):
    """Apply preprocessing steps to improve face recognition."""
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)  # Convert the image to grayscale
    equalized = cv2.equalizeHist(gray)  # Apply histogram equalization to improve contrast
    blurred = cv2.GaussianBlur(equalized, (5, 5), 0)  # Apply Gaussian blur to reduce noise
    return cv2.cvtColor(blurred, cv2.COLOR_GRAY2BGR)  # Convert the image back to BGR color space

def find_face_encodings(images):
    """Find face encodings for a list of images."""
    encode_list = []  # Initialize an empty list to store the face encodings
    for img in images:  # Iterate over each image in the list
        img = preprocess_image(img)  # Apply preprocessing to the image
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)  # Convert the image from BGR to RGB color space (required by face_recognition)
        try:
            encode = face_recognition.face_encodings(img)[0]  # Generate the face encoding for the image
            encode_list.append(encode)  # Add the face encoding to the list
        except IndexError:  # Handle the case where no face is found in the image
            print(f"No face found in one of the training images.")
    return encode_list  # Return the list of face encodings

def create_attendance_file(file_path):
    """Create a new attendance Excel file with headers."""
    workbook = openpyxl.Workbook()  # Create a new Excel workbook
    sheet = workbook.active  # Get the active worksheet
    sheet.append(['Name', 'Date', 'Time'])  # Add the headers to the worksheet
    workbook.save(file_path)  # Save the workbook to the specified file path

def mark_attendance(name, file_path):
    """Mark attendance in the Excel file."""
    try:
        workbook = openpyxl.load_workbook(file_path)  # Load the existing Excel workbook
        sheet = workbook.active  # Get the active worksheet
    except FileNotFoundError:  # Handle the case where the Excel file does not exist
        create_attendance_file(file_path)  # Create a new Excel file
        workbook = openpyxl.load_workbook(file_path)  # Load the newly created Excel workbook
        sheet = workbook.active  # Get the active worksheet

    now = datetime.now()  # Get the current date and time
    today_date = now.strftime("%Y-%m-%d")  # Format the date as YYYY-MM-DD
    current_time = now.strftime("%H:%M:%S")  # Format the time as HH:MM:SS

    # Check if the name already exists for today's date
    for row in range(2, sheet.max_row + 1):  # Iterate over each row in the worksheet (starting from row 2 to skip headers)
        existing_name = sheet.cell(row=row, column=1).value  # Get the value in the 'Name' column
        existing_date = sheet.cell(row=row, column=2).value  # Get the value in the 'Date' column

        if existing_name == name and existing_date == today_date:  # Check if the name and date match the current entry
            print(f'{name} already marked for attendance today.')
            return  # If they match, exit the function

    # If the name doesn't exist for today, add a new row
    sheet.append([name, today_date, current_time])  # Add a new row with the name, date, and time
    workbook.save(file_path)  # Save the changes to the Excel file
    print(f'Attendance marked for {name} at {current_time}')

def main():
    # Load training images and find encodings
    images, class_names = load_training_images(TRAINING_IMAGES_PATH)  # Load the training images and class names
    encode_list_known = find_face_encodings(images)  # Generate face encodings for the training images
    print('Encoding Complete')

    # Initialize webcam
    cap = cv2.VideoCapture(0)  # Open the default webcam (camera index 0)
    if not cap.isOpened():  # Check if the webcam was successfully opened
        print("Error: Could not open webcam.")
        return  # If not, exit the function

    while True:  # Start an infinite loop to continuously capture frames from the webcam
        success, img = cap.read()  # Read a frame from the webcam
        if not success:  # Check if a frame was successfully read
            print("Error: Could not read frame.")
            break  # If not, exit the loop

        preprocessed_img = preprocess_image(img)  # Preprocess the image to improve face recognition
        img_small = cv2.resize(preprocessed_img, (0, 0), None, 0.25, 0.25)  # Resize the image to 1/4 of its original size to speed up processing
        img_small = cv2.cvtColor(img_small, cv2.COLOR_BGR2RGB)  # Convert the image from BGR to RGB color space

        faces_cur_frame = face_recognition.face_locations(img_small)  # Find the locations of faces in the current frame
        encodes_cur_frame = face_recognition.face_encodings(img_small, faces_cur_frame)  # Generate face encodings for the faces in the current frame

        for encode_face, face_loc in zip(encodes_cur_frame, faces_cur_frame):  # Iterate over each face encoding and its corresponding location
            matches = face_recognition.compare_faces(encode_list_known, encode_face)  # Compare the current face encoding with the known face encodings
            face_dis = face_recognition.face_distance(encode_list_known, encode_face)  # Calculate the distance between the current face encoding and the known face encodings
            match_index = np.argmin(face_dis)  # Find the index of the closest match

            if matches[match_index] and face_dis[match_index] < FACE_RECOGNITION_THRESHOLD:  # Check if the closest match is below the specified threshold
                name = class_names[match_index].upper()  # Get the name of the matched person and convert it to uppercase
                y1, x2, y2, x1 = face_loc  # Extract the coordinates of the face location
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4  # Scale the coordinates back to the original image size
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)  # Draw a rectangle around the face
                cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)  # Draw a filled rectangle below the face for the name
                cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)  # Add the name to the image
                mark_attendance(name, ATTENDANCE_EXCEL_FILE)  # Mark the attendance of the person
            else:
                print("No Match")  # If no match is found, print "No Match"

        cv2.imshow('Webcam', img)  # Display the image with the face detections
        if cv2.waitKey(1) & 0xFF == ord('q'):  # Wait for a key press (1ms) and check if the pressed key is 'q'
            break  # If 'q' is pressed, exit the loop

    cap.release()  # Release the webcam
    cv2.destroyAllWindows()  # Close all OpenCV windows

if __name__ == "__main__":
    main()  # Call the main function if the script is executed directly
